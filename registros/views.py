from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import login
from django.http import HttpResponse
from .models import Registro
import openpyxl

@login_required
def lista(request):
    registros = Registro.objects.all()
    lotes = Registro.objects.values_list('lote', flat=True).distinct()
    sectores = Registro.objects.values_list('sector', flat=True).distinct()
    return render(request, 'registros/lista.html', {
        'registros': registros,
        'lotes': lotes,
        'sectores': sectores
    })

@login_required
def agregar(request):
    if request.method == 'POST':
        Registro.objects.create(
            numero=request.POST['numero'],
            codigo=request.POST['codigo'],
            chip=request.POST['chip'],
            sexo=request.POST['sexo'],
            raza=request.POST['raza'],
            color=request.POST['color'],
            detalle=request.POST.get('detalle', ''),
            lote=request.POST.get('lote', ''),
            sector=request.POST.get('sector', ''),
        )
        return redirect('lista')
    return render(request, 'registros/agregar.html')

@login_required
def eliminar(request, id):
    registro = get_object_or_404(Registro, id=id)
    registro.delete()
    return redirect('lista')

@login_required
def editar(request, id):
    registro = get_object_or_404(Registro, id=id)
    if request.method == 'POST':
        registro.numero = request.POST['numero']
        registro.codigo = request.POST['codigo']
        registro.chip = request.POST['chip']
        registro.sexo = request.POST['sexo']
        registro.raza = request.POST['raza']
        registro.color = request.POST['color']
        registro.detalle = request.POST.get('detalle', '')
        registro.lote = request.POST.get('lote', '')
        registro.sector = request.POST.get('sector', '')
        registro.save()
        return redirect('lista')
    return render(request, 'registros/editar.html', {'registro': registro})

def registro_usuario(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('lista')
    else:
        form = UserCreationForm()
    return render(request, 'registration/registro.html', {'form': form})

def descargar_plantilla(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla"
    ws.append(['CODIGO', 'CHIP', 'SEXO', 'RAZA', 'COLOR'])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=plantilla_ganado.xlsx'
    wb.save(response)
    return response

@login_required
def importar_excel(request):
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        lote = request.POST.get('lote', '')
        sector = request.POST.get('sector', '')
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        filas_importadas = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                Registro.objects.create(
                    codigo=row[0] or '',
                    chip=row[1] or '',
                    sexo=row[2] or '',
                    raza=row[3] or '',
                    color=row[4] or '',
                    lote=lote,
                    sector=sector,
                    detalle='',
                )
                filas_importadas += 1
        from django.contrib import messages
        messages.success(request, f'Se importaron {filas_importadas} registros correctamente.')
        return redirect('lista')
    return render(request, 'registros/importar.html')

@login_required
def eliminar_masivo(request):
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        Registro.objects.filter(id__in=ids).delete()
    return redirect('lista')

@login_required
def dashboard(request):
    from django.db.models import Count
    
    total = Registro.objects.count()
    
    por_raza = list(Registro.objects.values('raza').annotate(total=Count('id')).order_by('-total'))
    por_lote = list(Registro.objects.values('lote').annotate(total=Count('id')).order_by('-total'))
    por_sector = list(Registro.objects.values('sector').annotate(total=Count('id')).order_by('-total'))
    por_sexo = list(Registro.objects.values('sexo').annotate(total=Count('id')).order_by('-total'))
    
    return render(request, 'registros/dashboard.html', {
        'total': total,
        'por_raza': por_raza,
        'por_lote': por_lote,
        'por_sector': por_sector,
        'por_sexo': por_sexo,
    })